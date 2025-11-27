#Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet. For example, when the program is run like this 
#> py MultiplicationTable.py 6 


import openpyxl, os, sys
from pathlib import Path
from openpyxl.styles import Font
#open WB
wb = openpyxl.Workbook()
sheet = wb['Sheet']
bold_font = Font(bold=True)
MULTIPLIER = 1
if len(sys.argv) > 1:
    MULTIPLIER = int(sys.argv[1])
top = []
side = []
print(f'Creating Multiplication Table as Excel Workbook...')
for i in range (MULTIPLIER + 1):
    #get cell objects
    top_row = sheet.cell(row=1, column=i+1)
    side_col = sheet.cell(row=i+1, column=1)
    #make it bold
    top_row.font = bold_font
    side_col.font = bold_font  
    #assign value to the objects
    sheet[top_row.coordinate] , sheet[side_col.coordinate] = i , i
    top_value = sheet[top_row.coordinate].value
    side_value = sheet[side_col.coordinate].value
    top.append(top_value)
    side.append(side_value)
for col_num in range(2, sheet.max_column + 1):
    for row_num in range(2, sheet.max_row + 1):
        sum_cell = sheet.cell(row=row_num , column=col_num)
        sheet[sum_cell.coordinate] = top[col_num-1] * side[row_num-1]

wb.save(f'MultTableOf_{MULTIPLIER}.xlsx')
print(f'... Saving as "MultTableOf_{MULTIPLIER}.xlsx", to your Current Working Directory')